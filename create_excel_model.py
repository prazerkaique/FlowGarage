#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Script para criar o novo modelo Excel de importação de veículos
com todos os campos e validações solicitados pelo usuário.
"""

from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def create_vehicle_import_excel():
    """Cria o modelo Excel para importação de veículos"""
    
    # Criar workbook
    wb = Workbook()
    
    # Remover sheet padrão e criar as sheets necessárias
    wb.remove(wb.active)
    ws_vehicles = wb.create_sheet("Veículos")
    ws_instructions = wb.create_sheet("Instruções")
    
    # === CONFIGURAÇÃO DE ESTILOS ===
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    section_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_alignment = Alignment(horizontal='center', vertical='center')
    
    # === DEFINIR CABEÇALHOS ===
    headers = [
        # Informações Básicas
        "ID", "Categoria", "Marca", "Modelo", "Placa", "Ano Fabricação", 
        "Ano Modelo", "KM", "Preço", "Cor", "Status", "Descrição",
        
        # Especificações Técnicas
        "Tipo Carroceria", "Transmissão", "Combustível", "Motor", "Portas", "Direção",
        
        # Características Especiais
        "Blindado", "IPVA Pago", "Leilão", "Licenciamento em Dia",
        
        # Opcionais
        "Ar Condicionado", "Direção Hidráulica/Elétrica", "Vidros Elétricos", 
        "Travas Elétricas", "Alarme", "Som", "GPS", "Câmera de Ré", 
        "Sensor Estacionamento", "Airbags", "ABS", "Controle Estabilidade",
        "Controle Tração", "Piloto Automático", "Banco Couro", 
        "Rodas Liga Leve", "Teto Solar", "Faróis Neblina", "Bluetooth",
        "USB", "Entrada Auxiliar", "Controle Volante", "Retrovisor Elétrico",
        "Retrovisor Retrátil"
    ]
    
    # === CRIAR CABEÇALHOS NA PLANILHA ===
    for col, header in enumerate(headers, 1):
        cell = ws_vehicles.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment
        cell.border = border
        
        # Ajustar largura das colunas
        column_letter = get_column_letter(col)
        if header in ["Descrição"]:
            ws_vehicles.column_dimensions[column_letter].width = 40
        elif header in ["Modelo", "Placa"]:
            ws_vehicles.column_dimensions[column_letter].width = 20
        elif header in ["Marca", "Cor", "Transmissão", "Combustível"]:
            ws_vehicles.column_dimensions[column_letter].width = 15
        else:
            ws_vehicles.column_dimensions[column_letter].width = 12
    
    # === CRIAR VALIDAÇÕES ===
    
    # Categoria (Carro ou Moto)
    categoria_validation = DataValidation(
        type="list",
        formula1='"Carro,Moto"',
        showDropDown=True
    )
    categoria_validation.add(f"B2:B1000")
    ws_vehicles.add_data_validation(categoria_validation)
    
    # Cores
    cores = "Branco,Preto,Prata,Cinza,Azul,Vermelho,Verde,Amarelo,Marrom,Bege,Dourado,Laranja,Rosa,Roxo,Outro"
    cor_validation = DataValidation(
        type="list",
        formula1=f'"{cores}"',
        showDropDown=True
    )
    cor_validation.add(f"J2:J1000")
    ws_vehicles.add_data_validation(cor_validation)
    
    # Status
    status_validation = DataValidation(
        type="list",
        formula1='"Disponível,Vendido,Reservado,Em Manutenção,Indisponível"',
        showDropDown=True
    )
    status_validation.add(f"K2:K1000")
    ws_vehicles.add_data_validation(status_validation)
    
    # Tipo de Carroceria
    carroceria_validation = DataValidation(
        type="list",
        formula1='"Sedan,Hatch,SUV,Pickup,Conversível,Coupé,Perua,Van,Minivan,Crossover,Outro"',
        showDropDown=True
    )
    carroceria_validation.add(f"M2:M1000")
    ws_vehicles.add_data_validation(carroceria_validation)
    
    # Transmissão
    transmissao_validation = DataValidation(
        type="list",
        formula1='"Manual,Automática,CVT,Automatizada,Semi-automática"',
        showDropDown=True
    )
    transmissao_validation.add(f"N2:N1000")
    ws_vehicles.add_data_validation(transmissao_validation)
    
    # Combustível
    combustivel_validation = DataValidation(
        type="list",
        formula1='"Flex,Gasolina,Etanol,Diesel,Elétrico,Híbrido,GNV"',
        showDropDown=True
    )
    combustivel_validation.add(f"O2:O1000")
    ws_vehicles.add_data_validation(combustivel_validation)
    
    # Motor
    motor_validation = DataValidation(
        type="list",
        formula1='"1.0,1.2,1.3,1.4,1.5,1.6,1.7,1.8,1.9,2.0 - 2.9,3.0 - 3.9,4.0 ou mais"',
        showDropDown=True
    )
    motor_validation.add(f"P2:P1000")
    ws_vehicles.add_data_validation(motor_validation)
    
    # Direção
    direcao_validation = DataValidation(
        type="list",
        formula1='"Mecânica,Hidráulica,Elétrica,Eletro-hidráulica"',
        showDropDown=True
    )
    direcao_validation.add(f"Q2:Q1000")
    ws_vehicles.add_data_validation(direcao_validation)
    
    # Validações Sim/Não para características especiais e opcionais
    sim_nao_validation = DataValidation(
        type="list",
        formula1='"Sim,Não"',
        showDropDown=True
    )
    
    # Aplicar validação Sim/Não para todas as características especiais e opcionais
    sim_nao_columns = list(range(18, len(headers) + 1))  # Colunas R até o final
    for col in sim_nao_columns:
        column_letter = get_column_letter(col)
        sim_nao_validation.add(f"{column_letter}2:{column_letter}1000")
    
    ws_vehicles.add_data_validation(sim_nao_validation)
    
    # === ADICIONAR LINHA DE EXEMPLO ===
    example_row = [
        "001",  # ID
        "Carro",  # Categoria
        "Citroën",  # Marca
        "C3 GLX 1.4",  # Modelo
        "ABC-1234",  # Placa
        "2020",  # Ano Fabricação
        "2021",  # Ano Modelo
        "25.000",  # KM
        "R$ 45.000,00",  # Preço
        "Branco",  # Cor
        "Disponível",  # Status
        "Veículo em excelente estado, único dono, todas as revisões em dia",  # Descrição
        "Hatch",  # Tipo Carroceria
        "Manual",  # Transmissão
        "Flex",  # Combustível
        "4",  # Portas
        "Hidráulica",  # Direção
        "Não",  # Blindado
        "Sim",  # IPVA Pago
        "Não",  # Leilão
        "Sim",  # Licenciamento em Dia
        "Sim",  # Ar Condicionado
        "Sim",  # Direção Hidráulica/Elétrica
        "Sim",  # Vidros Elétricos
        "Sim",  # Travas Elétricas
        "Sim",  # Alarme
        "Sim",  # Som
        "Não",  # GPS
        "Não",  # Câmera de Ré
        "Não",  # Sensor Estacionamento
        "Sim",  # Airbags
        "Sim",  # ABS
        "Não",  # Controle Estabilidade
        "Não",  # Controle Tração
        "Não",  # Piloto Automático
        "Não",  # Banco Couro
        "Não",  # Rodas Liga Leve
        "Não",  # Teto Solar
        "Sim",  # Faróis Neblina
        "Não",  # Bluetooth
        "Não",  # USB
        "Sim",  # Entrada Auxiliar
        "Não",  # Controle Volante
        "Sim",  # Retrovisor Elétrico
        "Não"   # Retrovisor Retrátil
    ]
    
    for col, value in enumerate(example_row, 1):
        ws_vehicles.cell(row=2, column=col, value=value)
    
    # === CRIAR SHEET DE INSTRUÇÕES ===
    instructions = [
        "INSTRUÇÕES PARA PREENCHIMENTO DO MODELO DE IMPORTAÇÃO DE VEÍCULOS",
        "",
        "CAMPOS OBRIGATÓRIOS:",
        "• ID: Identificador único do veículo (ex: 001, 002, 003...)",
        "• Categoria: Selecione 'Carro' ou 'Moto'",
        "• Marca: Nome da marca do veículo",
        "• Modelo: Descrição completa do modelo (ex: C3 GLX 1.4, Civic EXL 2.0)",
        "• Placa: Placa do veículo no formato ABC-1234",
        "• Ano Fabricação: Ano de fabricação do veículo",
        "• Ano Modelo: Ano do modelo do veículo",
        "• KM: Quilometragem separada por ponto (ex: 10.000, 25.500)",
        "• Preço: Valor com R$ e separado por ponto (ex: R$ 25.000,00)",
        "• Cor: Selecione uma cor da lista disponível",
        "• Status: Selecione o status atual do veículo",
        "",
        "ESPECIFICAÇÕES TÉCNICAS:",
        "• Tipo Carroceria: Selecione o tipo de carroceria",
        "• Transmissão: Tipo de câmbio do veículo",
        "• Combustível: Tipo de combustível",
        "• Portas: Número de portas (ex: 2, 4, 5)",
        "• Direção: Tipo de direção do veículo",
        "",
        "CARACTERÍSTICAS ESPECIAIS:",
        "• Marque 'Sim' ou 'Não' para cada característica",
        "• Blindado: Se o veículo possui blindagem",
        "• IPVA Pago: Se o IPVA está em dia",
        "• Leilão: Se o veículo é proveniente de leilão",
        "• Licenciamento em Dia: Se o licenciamento está regular",
        "",
        "OPCIONAIS:",
        "• Marque 'Sim' ou 'Não' para cada opcional disponível",
        "• Lista completa de opcionais disponíveis na planilha",
        "",
        "OBSERVAÇÕES IMPORTANTES:",
        "• Preencha todos os campos obrigatórios",
        "• Use as listas de seleção quando disponíveis",
        "• Mantenha a formatação dos campos de KM e Preço",
        "• A descrição pode conter informações adicionais sobre o veículo",
        "",
        "ESTRUTURA DE PASTAS PARA MÍDIAS:",
        "• Crie uma pasta 'Mídia' no ZIP",
        "• Dentro de 'Mídia', crie uma pasta para cada ID de veículo",
        "• Dentro da pasta do ID, organize em: Fotos/, Vídeos/, Laudo/",
        "• Exemplo: Mídia/001/Fotos/, Mídia/001/Vídeos/, Mídia/001/Laudo/"
    ]
    
    for row, instruction in enumerate(instructions, 1):
        cell = ws_instructions.cell(row=row, column=1, value=instruction)
        if row == 1:
            cell.font = Font(bold=True, size=14)
        elif instruction.startswith("•"):
            cell.font = Font(size=10)
        else:
            cell.font = Font(bold=True, size=11)
    
    # Ajustar largura da coluna de instruções
    ws_instructions.column_dimensions['A'].width = 80
    
    # === SALVAR ARQUIVO ===
    output_path = "/Users/kaiquemoreira/Documents/IA de Carro/frontend/public/modelo-importacao-veiculos.xlsx"
    wb.save(output_path)
    print(f"✅ Novo modelo Excel criado com sucesso: {output_path}")
    print(f"📊 Total de campos: {len(headers)}")
    print("🎯 Recursos incluídos:")
    print("   • Validações de dados com listas de seleção")
    print("   • Formatação profissional")
    print("   • Exemplo preenchido")
    print("   • Instruções detalhadas")
    print("   • Estrutura completa de campos solicitados")

if __name__ == "__main__":
    create_vehicle_import_excel()